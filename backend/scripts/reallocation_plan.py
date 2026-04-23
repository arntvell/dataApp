"""
SALE_FW25 Stock Reallocation Plan Generator

Reads variant-level CSV data and produces an actionable XLSX transfer plan
for redistributing stock from low-performing to sold-out locations.

Usage:
    docker compose exec backend python scripts/reallocation_plan.py
    docker compose cp backend:/tmp/sale_fw25_reallocation.xlsx ~/Desktop/
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

CSV_PATH = "/tmp/sale_fw25_variants.csv"
OUTPUT_PATH = "/tmp/sale_fw25_reallocation.xlsx"

# Physical stores only (exclude Online)
STORES = ["Livid Oslo", "Livid Trondheim", "Livid Bergen", "Livid Stavanger", "Past Løkka"]

# Location sell-through rates for secondary tiebreaker (from analysis)
LOCATION_PRIORITY = {
    "Livid Oslo": 5,
    "Livid Bergen": 4,
    "Livid Trondheim": 3,
    "Livid Stavanger": 2,
    "Past Løkka": 1,
}

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Route colors for visual grouping
ROUTE_COLORS = {
    "Livid Oslo": "E2EFDA",
    "Livid Bergen": "D6E4F0",
    "Livid Trondheim": "FCE4D6",
    "Livid Stavanger": "E2D9F3",
    "Past Løkka": "FFF2CC",
}


def load_data(path: str) -> pd.DataFrame:
    """Load and parse the variant CSV."""
    df = pd.read_csv(path, sep=";")
    # Rename columns for easier access
    df = df.rename(columns={
        "Product": "product",
        "Size/Variant": "size",
        "SKU": "sku",
    })
    # Build per-store sold/stock columns into a normalized structure
    records = []
    for _, row in df.iterrows():
        for store in STORES:
            sold = int(row.get(f"{store} Sold", 0))
            stock = int(row.get(f"{store} Stock", 0))
            records.append({
                "product": row["product"],
                "size": str(row["size"]),
                "sku": row["sku"],
                "store": store,
                "sold": sold,
                "stock": max(0, stock),  # clamp negative stock to 0
            })
    return pd.DataFrame(records)


# If a donor has this many or fewer total units for a product family,
# consolidate everything to the top performer (one shipment).
# Above this threshold, distribute across needy performers per size.
CONSOLIDATION_THRESHOLD = 3


def run_allocation(df: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """
    Performance-based reallocation algorithm.

    For each product family:
      - Performers: stores with sales on this family
      - Donors: stores with zero sales on this family
      - Needy: performers that are out of stock on a specific size

    Allocation rules:
      1. Replenish performers that lack stock, prioritized by family sales
      2. If multiple performers need stock AND donor has enough → split
      3. If donor has low total family stock → consolidate to top performer
         (avoids creating too many small shipments)

    Returns (transfers, cannot_fulfill).
    """
    transfers = []
    cannot_fulfill = []

    # --- Step 1: Compute family-level sales per store ---
    family_sales = (
        df.groupby(["product", "store"])["sold"]
        .sum()
        .reset_index()
        .rename(columns={"sold": "family_sold"})
    )

    family_sales_map = {}
    for _, row in family_sales.iterrows():
        family_sales_map[(row["product"], row["store"])] = row["family_sold"]

    # --- Step 2: For each product, classify stores ---
    products = df["product"].unique()

    for product in products:
        product_df = df[df["product"] == product]
        stores_in_play = product_df["store"].unique()

        performers = []
        non_performers = []
        for store in stores_in_play:
            fam_sold = family_sales_map.get((product, store), 0)
            if fam_sold > 0:
                performers.append((store, fam_sold))
            else:
                non_performers.append(store)

        if not performers:
            continue

        # Sort performers: most sales first, tiebreak by location priority
        performers.sort(key=lambda x: (-x[1], -LOCATION_PRIORITY.get(x[0], 0)))
        top_performer = performers[0][0]
        total_performer_sales = sum(s for _, s in performers)

        # Build per-size store data
        variants = product_df.groupby(["size", "sku"])
        variant_store_data = {}
        for (size, sku), var_group in variants:
            store_data = {}
            for _, row in var_group.iterrows():
                store_data[row["store"]] = {
                    "sold": row["sold"],
                    "stock": row["stock"],
                }
            variant_store_data[(size, sku)] = store_data

        # --- Step 3: Process each donor separately ---
        # Each donor decides: consolidate or distribute?
        for donor_store in non_performers:
            # Calculate this donor's total family stock
            donor_family_stock = 0
            donor_sizes = []  # (size, sku, qty)
            for (size, sku), store_data in variant_store_data.items():
                sd = store_data.get(donor_store, {"sold": 0, "stock": 0})
                if sd["stock"] > 0:
                    donor_sizes.append((size, sku, sd["stock"]))
                    donor_family_stock += sd["stock"]

            if donor_family_stock == 0:
                continue

            # Decision: consolidate or distribute?
            consolidate = donor_family_stock <= CONSOLIDATION_THRESHOLD

            if consolidate:
                # Send everything to the top performer that needs stock
                # (or just the overall top performer)
                destination = top_performer
                for store, fam_sold in performers:
                    # Find the highest-ranked performer that is out of stock
                    # on any of the sizes this donor has
                    for (size, sku, qty) in donor_sizes:
                        sd = variant_store_data[(size, sku)].get(
                            store, {"sold": 0, "stock": 0}
                        )
                        if sd["stock"] == 0:
                            destination = store
                            break
                    if destination != top_performer:
                        break

                for (size, sku, qty) in donor_sizes:
                    sd_recv = variant_store_data[(size, sku)].get(
                        destination, {"sold": 0, "stock": 0}
                    )
                    transfers.append({
                        "from_store": donor_store,
                        "to_store": destination,
                        "product": product,
                        "size": size,
                        "sku": sku,
                        "qty": qty,
                        "reason": (
                            f"Consolidated to {destination} "
                            f"(family: {family_sales_map.get((product, destination), 0)} sold). "
                            f"{donor_store} has {donor_family_stock} total, sold 0."
                        ),
                    })
            else:
                # Distribute: for each size, send to needy performers
                for (size, sku, donor_qty) in donor_sizes:
                    store_data = variant_store_data[(size, sku)]

                    # Find needy performers: out of stock on THIS size
                    needy = [
                        (store, fam_sold)
                        for store, fam_sold in performers
                        if store_data.get(store, {"sold": 0, "stock": 0})["stock"] == 0
                    ]

                    if not needy:
                        # No performer needs this size — send to top performer
                        # to build range depth
                        needy = [performers[0]]

                    if len(needy) == 1:
                        # One destination — send all
                        recv_store = needy[0][0]
                        sd_recv = store_data.get(recv_store, {"sold": 0, "stock": 0})
                        transfers.append({
                            "from_store": donor_store,
                            "to_store": recv_store,
                            "product": product,
                            "size": size,
                            "sku": sku,
                            "qty": donor_qty,
                            "reason": (
                                f"{recv_store} needs restock "
                                f"(family: {family_sales_map.get((product, recv_store), 0)} sold). "
                                f"{donor_store} has {donor_qty}, sold 0."
                            ),
                        })
                    else:
                        # Multiple needy — split proportionally by family sales
                        needy_total_sales = sum(s for _, s in needy)
                        allocation = {}
                        for store, fam_sold in needy:
                            share = fam_sold / needy_total_sales
                            units = int(donor_qty * share)
                            if units > 0:
                                allocation[store] = units

                        # Leftover from rounding → top needy performer
                        allocated = sum(allocation.values())
                        leftover = donor_qty - allocated
                        for store, fam_sold in needy:
                            if leftover <= 0:
                                break
                            allocation[store] = allocation.get(store, 0) + 1
                            leftover -= 1

                        for recv_store, qty in allocation.items():
                            if qty <= 0:
                                continue
                            sd_recv = store_data.get(recv_store, {"sold": 0, "stock": 0})
                            transfers.append({
                                "from_store": donor_store,
                                "to_store": recv_store,
                                "product": product,
                                "size": size,
                                "sku": sku,
                                "qty": qty,
                                "reason": (
                                    f"{recv_store} needs restock "
                                    f"(family: {family_sales_map.get((product, recv_store), 0)} sold). "
                                    f"Split from {donor_store} ({donor_qty} avail)."
                                ),
                            })

        # --- Record cannot-fulfill for needy performers with no donors ---
        for (size, sku), store_data in variant_store_data.items():
            has_any_donor = any(
                store_data.get(d, {"sold": 0, "stock": 0})["stock"] > 0
                for d in non_performers
            )
            if has_any_donor:
                continue
            for store, fam_sold in performers:
                sd = store_data.get(store, {"sold": 0, "stock": 0})
                if sd["sold"] > 0 and sd["stock"] == 0:
                    cannot_fulfill.append({
                        "product": product,
                        "size": size,
                        "sku": sku,
                        "store": store,
                        "units_sold": sd["sold"],
                        "reason": "No non-performer store holds stock for this variant",
                    })

    # Sort: destination (Oslo first), then product, then size
    transfers.sort(
        key=lambda x: (-LOCATION_PRIORITY.get(x["to_store"], 0), x["product"], x["size"])
    )

    return transfers, cannot_fulfill


def build_summary_by_route(transfers: list[dict]) -> pd.DataFrame:
    """Aggregate transfers by route (from → to)."""
    if not transfers:
        return pd.DataFrame(columns=["From", "To", "Total SKUs", "Total Units"])

    df = pd.DataFrame(transfers)
    summary = (
        df.groupby(["from_store", "to_store"])
        .agg(total_skus=("sku", "count"), total_units=("qty", "sum"))
        .reset_index()
        .rename(columns={
            "from_store": "From",
            "to_store": "To",
            "total_skus": "Total SKUs",
            "total_units": "Total Units",
        })
        .sort_values(["To", "From"])
    )
    return summary


def build_store_impact(df_raw: pd.DataFrame, transfers: list[dict]) -> pd.DataFrame:
    """Calculate per-store impact: current stock, sending, receiving, net, new stock."""
    # Current stock per store
    store_stock = df_raw.groupby("store")["stock"].sum().to_dict()

    sending = {}
    receiving = {}
    for t in transfers:
        sending[t["from_store"]] = sending.get(t["from_store"], 0) + t["qty"]
        receiving[t["to_store"]] = receiving.get(t["to_store"], 0) + t["qty"]

    rows = []
    for store in STORES:
        current = store_stock.get(store, 0)
        out = sending.get(store, 0)
        inc = receiving.get(store, 0)
        net = inc - out
        rows.append({
            "Location": store,
            "Current Stock": current,
            "Units Sending": out,
            "Units Receiving": inc,
            "Net Change": net,
            "New Stock Level": current + net,
        })
    return pd.DataFrame(rows)


def style_sheet(ws, num_cols: int, header_row: int = 1):
    """Apply consistent styling to a worksheet."""
    # Style headers
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Auto-width columns
    for col in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze top row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Light borders on data rows
    for row in ws.iter_rows(min_row=header_row + 1, max_col=num_cols):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def write_xlsx(
    transfers: list[dict],
    cannot_fulfill: list[dict],
    df_raw: pd.DataFrame,
):
    """Write the 4-sheet XLSX workbook."""
    wb = Workbook()

    # --- Sheet 1: Transfer Plan ---
    ws1 = wb.active
    ws1.title = "Transfer Plan"
    headers1 = ["From", "To", "Product", "Size", "SKU", "Qty", "Reason"]
    ws1.append(headers1)

    for t in transfers:
        ws1.append([
            t["from_store"],
            t["to_store"],
            t["product"],
            t["size"],
            t["sku"],
            t["qty"],
            t["reason"],
        ])

    # Color-code rows by destination
    for row_idx in range(2, len(transfers) + 2):
        dest = ws1.cell(row=row_idx, column=2).value
        if dest in ROUTE_COLORS:
            fill = PatternFill(
                start_color=ROUTE_COLORS[dest],
                end_color=ROUTE_COLORS[dest],
                fill_type="solid",
            )
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=row_idx, column=col).fill = fill

    style_sheet(ws1, len(headers1))

    # --- Sheet 2: Summary by Route ---
    ws2 = wb.create_sheet("Summary by Route")
    route_df = build_summary_by_route(transfers)
    headers2 = list(route_df.columns)
    ws2.append(headers2)
    for _, row in route_df.iterrows():
        ws2.append(list(row))

    # Add totals row
    if len(route_df) > 0:
        total_row = len(route_df) + 2
        ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws2.cell(row=total_row, column=3, value=route_df["Total SKUs"].sum()).font = Font(bold=True)
        ws2.cell(row=total_row, column=4, value=route_df["Total Units"].sum()).font = Font(bold=True)

    style_sheet(ws2, len(headers2))

    # --- Sheet 3: Store Impact ---
    ws3 = wb.create_sheet("Store Impact")
    impact_df = build_store_impact(df_raw, transfers)
    headers3 = list(impact_df.columns)
    ws3.append(headers3)
    for _, row in impact_df.iterrows():
        ws3.append(list(row))

    # Color net-positive green, net-negative red
    for row_idx in range(2, len(impact_df) + 2):
        net_cell = ws3.cell(row=row_idx, column=5)  # Net Change column
        if net_cell.value and net_cell.value > 0:
            net_cell.font = Font(color="006100")
            net_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif net_cell.value and net_cell.value < 0:
            net_cell.font = Font(color="9C0006")
            net_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    style_sheet(ws3, len(headers3))

    # --- Sheet 4: Cannot Fulfill ---
    ws4 = wb.create_sheet("Cannot Fulfill")
    headers4 = ["Product", "Size", "SKU", "Store (Needs Stock)", "Units Sold", "Reason"]
    ws4.append(headers4)

    # Sort: most sold first
    cannot_fulfill.sort(key=lambda x: (-x["units_sold"], x["product"], x["size"]))

    for cf in cannot_fulfill:
        ws4.append([
            cf["product"],
            cf["size"],
            cf["sku"],
            cf["store"],
            cf["units_sold"],
            cf["reason"],
        ])

    style_sheet(ws4, len(headers4))

    # --- Per-store sending sheets ---
    # Group transfers by sending store
    sending_by_store = {}
    for t in transfers:
        sending_by_store.setdefault(t["from_store"], []).append(t)

    send_headers = ["To", "Product", "Size", "SKU", "Qty"]

    for store in STORES:
        store_transfers = sending_by_store.get(store, [])
        if not store_transfers:
            continue

        # Short sheet name
        short = store.replace("Livid ", "").replace("Past ", "Past ")
        ws = wb.create_sheet(f"Send - {short}")

        ws.append(send_headers)

        # Sort by destination, then product, then size
        store_transfers.sort(
            key=lambda x: (
                -LOCATION_PRIORITY.get(x["to_store"], 0),
                x["product"],
                x["size"],
            )
        )

        for t in store_transfers:
            ws.append([
                t["to_store"],
                t["product"],
                t["size"],
                t["sku"],
                t["qty"],
            ])

        # Color-code by destination
        for row_idx in range(2, len(store_transfers) + 2):
            dest = ws.cell(row=row_idx, column=1).value
            if dest in ROUTE_COLORS:
                fill = PatternFill(
                    start_color=ROUTE_COLORS[dest],
                    end_color=ROUTE_COLORS[dest],
                    fill_type="solid",
                )
                for col in range(1, len(send_headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

        # Totals row
        total_row = len(store_transfers) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        total_qty = sum(t["qty"] for t in store_transfers)
        ws.cell(row=total_row, column=5, value=total_qty).font = Font(bold=True)
        total_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        for col in range(1, len(send_headers) + 1):
            ws.cell(row=total_row, column=col).fill = total_fill

        style_sheet(ws, len(send_headers))

    wb.save(OUTPUT_PATH)


def main():
    print("Loading variant data...")
    df = load_data(CSV_PATH)
    print(f"  {len(df)} variant-location records loaded")

    # Quick stats
    physical = df[df["store"].isin(STORES)]
    total_sold = physical["sold"].sum()
    total_stock = physical["stock"].sum()
    print(f"  Physical stores: {total_sold} sold, {total_stock} stock remaining")

    print("\nRunning allocation algorithm...")
    transfers, cannot_fulfill = run_allocation(df)

    total_units = sum(t["qty"] for t in transfers)
    print(f"  {len(transfers)} transfer actions generated ({total_units} total units)")
    print(f"  {len(cannot_fulfill)} variants cannot be fulfilled")

    # Verification
    print("\nVerification:")
    # Check no negative stock
    donor_usage = {}
    for t in transfers:
        key = (t["from_store"], t["sku"])
        donor_usage[key] = donor_usage.get(key, 0) + t["qty"]

    original_stock = {}
    for _, row in df.iterrows():
        key = (row["store"], row["sku"])
        original_stock[key] = row["stock"]

    violations = 0
    for key, used in donor_usage.items():
        orig = original_stock.get(key, 0)
        if used > orig:
            print(f"  WARNING: {key} over-allocated: used {used}, had {orig}")
            violations += 1

    if violations == 0:
        print("  OK: No stock violations")

    # Destination summary
    dest_totals = {}
    for t in transfers:
        dest_totals[t["to_store"]] = dest_totals.get(t["to_store"], 0) + t["qty"]
    source_totals = {}
    for t in transfers:
        source_totals[t["from_store"]] = source_totals.get(t["from_store"], 0) + t["qty"]

    print("\n  Receiving:")
    for store in STORES:
        if store in dest_totals:
            print(f"    {store}: +{dest_totals[store]} units")

    print("\n  Sending:")
    for store in STORES:
        if store in source_totals:
            print(f"    {store}: -{source_totals[store]} units")

    print(f"\nWriting XLSX to {OUTPUT_PATH}...")
    write_xlsx(transfers, cannot_fulfill, df)
    print("Done!")


if __name__ == "__main__":
    main()
