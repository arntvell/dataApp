"""
SALE_FW25 Store Report — XLSX with one sheet per physical store.

Each sheet shows all sale products with: Items Sold, Revenue, Stock.
Sorted by revenue descending. Includes totals row and conditional formatting.

Usage:
    docker compose exec backend python scripts/sale_store_report.py
"""

import csv
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import settings
from connectors.shopify_connector import ShopifyConnector
from database.config import SessionLocal
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/tmp/sale_fw25_store_report.xlsx"
PLANNER_PATH = "/tmp/sale_planner_input.csv"
SALE_TAG = "SALE_FW25"
SALE_START = "2025-12-27"

PHYSICAL_STORES = [
    "Livid Oslo",
    "Livid Bergen",
    "Livid Trondheim",
    "Livid Stavanger",
    "Past Løkka",
]

SENTRALLAGER = "Livid Sentrallager"

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
SOLD_OUT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SOLD_OUT_FONT = Font(color="9C0006")
HIGH_SELL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HIGH_SELL_FONT = Font(color="006100")


def read_planner() -> dict[str, dict]:
    """Read planner CSV for brand/category info. Returns {family: {brand, category}}."""
    families = {}
    with open(PLANNER_PATH, "r", newline="") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) >= 3 and row[2].strip():
                families[row[2].strip()] = {
                    "brand": row[0].strip(),
                    "category": row[1].strip(),
                }
    return families


def fetch_shopify_products() -> dict[str, list[str]]:
    """Fetch SALE_FW25 products. Returns {title: [skus]}."""
    connector = ShopifyConnector(settings.get_connector_configs()["shopify"])
    products = {}
    cursor = None

    while True:
        after_clause = f', after: "{cursor}"' if cursor else ""
        query = f"""
        {{
            products(first: 50, query: "tag:{SALE_TAG}"{after_clause}) {{
                edges {{
                    cursor
                    node {{
                        title
                        variants(first: 100) {{
                            edges {{
                                node {{
                                    sku
                                }}
                            }}
                        }}
                    }}
                }}
                pageInfo {{
                    hasNextPage
                }}
            }}
        }}
        """
        result = connector._make_graphql_request(query)
        if "errors" in result and result["errors"]:
            break

        data = result.get("data", {}).get("products", {})
        edges = data.get("edges", [])

        for edge in edges:
            node = edge["node"]
            title = node["title"]
            skus = [
                v["node"]["sku"].strip()
                for v in node.get("variants", {}).get("edges", [])
                if v["node"].get("sku", "").strip()
            ]
            if skus:
                products.setdefault(title, []).extend(skus)
            cursor = edge["cursor"]

        if not data.get("pageInfo", {}).get("hasNextPage", False) or not edges:
            break

    return products


def match_sale_products(planner_info: dict[str, dict]) -> dict[str, list[str]]:
    """
    Use Shopify SALE_FW25 tag as the authoritative source of sale products.
    Planner provides brand/category metadata only.
    Returns {product_title: [skus]} for products tagged SALE_FW25.
    """
    shopify = fetch_shopify_products()
    print(f"  {len(shopify)} sale products from Shopify SALE_FW25 tag\n")
    return shopify


def query_store_data(all_skus: list[str]) -> tuple[dict, dict, dict]:
    """
    Returns:
      sales: {(sku, location): qty_sold}
      revenue: {(sku, location): revenue}
      stock: {(sku, location): available}
    """
    db = SessionLocal()
    try:
        # Sales + revenue
        q = text("""
            SELECT soi.sku, so.location,
                   SUM(soi.quantity) AS qty,
                   SUM(soi.line_total) AS rev
            FROM sales_order_items soi
            JOIN sales_orders so ON soi.order_id = so.id
            WHERE soi.sku = ANY(:skus)
              AND so.order_date >= :sale_start
              AND so.location IS NOT NULL
            GROUP BY soi.sku, so.location
        """)
        result = db.execute(q, {"skus": all_skus, "sale_start": SALE_START})
        sales = {}
        revenue = {}
        for row in result:
            sales[(row.sku, row.location)] = int(row.qty)
            revenue[(row.sku, row.location)] = float(row.rev)

        # Stock
        all_locs = PHYSICAL_STORES + [SENTRALLAGER]
        q2 = text("""
            SELECT sku, location, available
            FROM raw.cin7_stock
            WHERE sku = ANY(:skus)
              AND location = ANY(:locations)
        """)
        result = db.execute(q2, {"skus": all_skus, "locations": all_locs})
        stock = {}
        for row in result:
            stock[(row.sku, row.location)] = int(row.available)

        return sales, revenue, stock
    finally:
        db.close()


def aggregate_for_store(
    family: str,
    skus: list[str],
    store: str,
    sales: dict,
    revenue: dict,
    stock: dict,
) -> dict:
    """Aggregate sold, revenue, stock for one family at one store."""
    total_sold = sum(sales.get((sku, store), 0) for sku in skus)
    total_rev = sum(revenue.get((sku, store), 0) for sku in skus)
    total_stock = sum(max(0, stock.get((sku, store), 0)) for sku in skus)

    if total_sold + total_stock > 0:
        sell_through = total_sold / (total_sold + total_stock)
    else:
        sell_through = None

    return {
        "sold": total_sold,
        "revenue": total_rev,
        "stock": total_stock,
        "sell_through": sell_through,
    }


def style_sheet(ws, num_cols, num_data_rows):
    """Apply styling to worksheet."""
    # Header styling
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx in range(2, num_data_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Revenue column formatting (col 5)
        rev_cell = ws.cell(row=row_idx, column=5)
        if rev_cell.value is not None:
            rev_cell.number_format = '#,##0'

        # Sell-through formatting (col 6)
        st_cell = ws.cell(row=row_idx, column=6)
        if st_cell.value is not None:
            st_cell.number_format = '0%'
            if st_cell.value >= 0.6:
                st_cell.fill = HIGH_SELL_FILL
                st_cell.font = HIGH_SELL_FONT
            elif st_cell.value == 0 and ws.cell(row=row_idx, column=4).value == 0:
                pass  # no sales, no stock — leave neutral

        # Stock = 0 but sold > 0: highlight sold-out
        stock_cell = ws.cell(row=row_idx, column=7)
        sold_cell = ws.cell(row=row_idx, column=4)
        if stock_cell.value == 0 and sold_cell.value and sold_cell.value > 0:
            stock_cell.fill = SOLD_OUT_FILL
            stock_cell.font = SOLD_OUT_FONT

    # Zebra striping
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_idx in range(2, num_data_rows + 2):
        if row_idx % 2 == 0:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col)
                # Don't override conditional fills
                if cell.fill == PatternFill():
                    cell.fill = zebra_fill

    # Auto-width
    for col in range(1, num_cols + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row_idx in range(2, min(num_data_rows + 2, 100)):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 45)

    # Freeze header
    ws.freeze_panes = "A2"


def main():
    print("Reading planner for product metadata...")
    planner_info = read_planner()
    family_names = list(planner_info.keys())
    print(f"  {len(family_names)} families\n")

    print("Fetching sale products (Shopify SALE_FW25 tag)...")
    matched = match_sale_products(planner_info)

    all_skus = list(set(sku for skus in matched.values() for sku in skus))
    print(f"Querying sales, revenue, and stock for {len(all_skus)} SKUs...")
    sales, revenue, stock = query_store_data(all_skus)
    print(f"  {len(sales)} sale records, {len(stock)} stock records\n")

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    headers = ["Brand", "Category", "Product", "Sold", "Revenue (NOK)", "Sell-Through", "Stock"]

    for store in PHYSICAL_STORES:
        # Short sheet name
        sheet_name = store.replace("Livid ", "").replace("Past ", "Past ")
        ws = wb.create_sheet(sheet_name)

        # Build rows for this store
        rows = []
        for family, skus in matched.items():
            info = planner_info.get(family, {"brand": "", "category": ""})
            agg = aggregate_for_store(family, skus, store, sales, revenue, stock)

            # Skip products with no activity and no stock at this store
            if agg["sold"] == 0 and agg["stock"] == 0:
                continue

            rows.append({
                "brand": info["brand"],
                "category": info["category"],
                "product": family,
                "sold": agg["sold"],
                "revenue": agg["revenue"],
                "sell_through": agg["sell_through"],
                "stock": agg["stock"],
            })

        # Sort by revenue desc
        rows.sort(key=lambda r: (-r["revenue"], -r["sold"], r["product"]))

        # Write header
        ws.append(headers)

        # Write data
        for r in rows:
            ws.append([
                r["brand"],
                r["category"],
                r["product"],
                r["sold"],
                r["revenue"],
                r["sell_through"],
                r["stock"],
            ])

        # Totals row
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=total_row, column=1).fill = TOTAL_FILL
        for col in range(2, len(headers) + 1):
            ws.cell(row=total_row, column=col).fill = TOTAL_FILL
            ws.cell(row=total_row, column=col).font = TOTAL_FONT

        total_sold = sum(r["sold"] for r in rows)
        total_rev = sum(r["revenue"] for r in rows)
        total_stock = sum(r["stock"] for r in rows)
        overall_st = total_sold / (total_sold + total_stock) if (total_sold + total_stock) > 0 else 0

        ws.cell(row=total_row, column=4, value=total_sold)
        ws.cell(row=total_row, column=5, value=total_rev).number_format = '#,##0'
        ws.cell(row=total_row, column=6, value=overall_st).number_format = '0%'
        ws.cell(row=total_row, column=7, value=total_stock)

        style_sheet(ws, len(headers), len(rows))

        print(f"  {sheet_name}: {len(rows)} products, {total_sold} sold, "
              f"{total_rev:,.0f} NOK, {total_stock} stock, {overall_st:.0%} sell-through")

    # Add summary sheet at the beginning
    ws_sum = wb.create_sheet("Overview", 0)
    sum_headers = ["Location", "Products", "Items Sold", "Revenue (NOK)", "Sell-Through", "Stock Remaining"]
    ws_sum.append(sum_headers)

    for store in PHYSICAL_STORES:
        sheet_name = store.replace("Livid ", "").replace("Past ", "Past ")
        ws_store = wb[sheet_name]
        total_row = ws_store.max_row
        products = ws_store.max_row - 2  # minus header and total row

        ws_sum.append([
            store,
            products,
            ws_store.cell(row=total_row, column=4).value,
            ws_store.cell(row=total_row, column=5).value,
            ws_store.cell(row=total_row, column=6).value,
            ws_store.cell(row=total_row, column=7).value,
        ])

    # Grand total
    grand_row = len(PHYSICAL_STORES) + 2
    ws_sum.cell(row=grand_row, column=1, value="TOTAL").font = TOTAL_FONT
    for col in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=grand_row, column=col).fill = TOTAL_FILL
        ws_sum.cell(row=grand_row, column=col).font = TOTAL_FONT

    grand_sold = sum(ws_sum.cell(row=r, column=3).value or 0 for r in range(2, grand_row))
    grand_rev = sum(ws_sum.cell(row=r, column=4).value or 0 for r in range(2, grand_row))
    grand_stock = sum(ws_sum.cell(row=r, column=6).value or 0 for r in range(2, grand_row))
    grand_st = grand_sold / (grand_sold + grand_stock) if (grand_sold + grand_stock) > 0 else 0

    ws_sum.cell(row=grand_row, column=3, value=grand_sold)
    ws_sum.cell(row=grand_row, column=4, value=grand_rev).number_format = '#,##0'
    ws_sum.cell(row=grand_row, column=5, value=grand_st).number_format = '0%'
    ws_sum.cell(row=grand_row, column=6, value=grand_stock)

    # Style overview
    for col in range(1, len(sum_headers) + 1):
        cell = ws_sum.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx in range(2, grand_row + 1):
        ws_sum.cell(row=row_idx, column=4).number_format = '#,##0'
        ws_sum.cell(row=row_idx, column=5).number_format = '0%'
        for col in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row_idx, column=col).border = THIN_BORDER
            ws_sum.cell(row=row_idx, column=col).alignment = Alignment(vertical="center")

    for col in range(1, len(sum_headers) + 1):
        max_len = max(
            len(str(ws_sum.cell(row=r, column=col).value or ""))
            for r in range(1, grand_row + 1)
        )
        ws_sum.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)

    ws_sum.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"\nSaved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
